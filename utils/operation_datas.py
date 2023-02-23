# _*_ coding: UTF-8 _*_
"""
@project -> file : city-test -> rr
@Author          : qinmin.vendor
@Date            : 2023/1/29 19:44
@Desc            : 操作excle和json的工具类
"""
import json
import os
import openpyxl
import xlsxwriter as xlsxwriter
from api_case_generate_tools.utils.exception_util import opeartionTypeError,sheetNameNotFoundError
from api_case_generate_tools.utils.other_util import otherUtil


class operationExcel(object):
    def __init__(self, file_path, file_name,opeartion_type='read',data_only=False,sheet_name='data'):
        self.file_path=file_path
        self.file_name=file_name
        self.sheet_name=sheet_name
        if self.file_path is None:
            self.file_path = os.path.join(otherUtil.get_project_rootpath(match_paths=['config','utils']), '../config')
        if self.file_name is None:
            self.file_name='DataCase_ALL.xlsx'
        self.file_address = os.path.join(self.file_path, self.file_name)
        if not os.path.exists(self.file_path):
            os.makedirs(self.file_path)

        if not opeartion_type in ('read','write','read_write'):
            raise opeartionTypeError(msg=f"opeartion_type错误,预期:'read','write','read_write',实际:{opeartion_type}")
        if opeartion_type in ('read','read_write'):
            self.check_file_address_exists(sheet_name=self.sheet_name)
            if opeartion_type=='read':
                read_only = True
            else:
                read_only = False
            self.workbook=openpyxl.load_workbook(self.file_address,data_only=data_only,read_only=read_only) #data_only=True,读取数据时保留exlce单元格格式
        elif opeartion_type=='write':
            self.workbook=openpyxl.Workbook()

        self.data = self.get_data()

    def check_file_address_exists(self,sheet_name):
        '''检查sheet是否存在，如果不存在则触发创建'''
        if not os.path.exists(self.file_address):
            # 创建工作薄
            new_book = openpyxl.Workbook()
            sheet = new_book.active
            sheet.title = sheet_name
            new_book.save(self.file_address)

    def create_sheet(self, title=None, index=None,is_save=False):
        '''创建sheet'''
        self.workbook.create_sheet( title=title, index=index)
        if is_save:
            self.save_workbook()

    def rename_sheet_name(self,src_sheet_name,target_sheet_name,is_save=False):
        '''
        重命名sheet_name
        Args:
            src_sheet_name: 需要修改名称的sheet_name
            target_sheet_name: 修改后的sheet_name
            is_save: 是否保存
        Returns:
        '''
        src_sheet_name_obj=self.get_data_for_sheet_name(sheet_name=src_sheet_name)
        src_sheet_name_obj.title=target_sheet_name
        if is_save:
            self.save_workbook()

    def copy_worksheet(self,src_sheet_name,target_sheet_name,is_save=False):
        '''
        复制工作簿内容
        Args:
            src_sheet_name:  需要被复制的sheet_name
            target_sheet_name: 复制后的sheet_name
            is_save: 是否保存
        Returns:
        '''
        src_sheet_name_obj=self.get_data_for_sheet_name(sheet_name=src_sheet_name)
        target_sheet_obj=self.workbook.copy_worksheet(src_sheet_name_obj)
        target_sheet_obj.title=target_sheet_name
        if is_save:
            self.save_workbook()


    def delete_sheet(self,sheet_name,is_save=False):
        '''删除sheet'''
        try:
            worksheet=self.get_data_for_sheet_name(sheet_name=sheet_name)
            self.workbook.remove(worksheet)
        except sheetNameNotFoundError:
            pass
        finally:
            if is_save:
                self.save_workbook()

    # 写入数据
    def write_value(self, row, col, value,is_save=False):
        '''写入excel数据row，col，value'''
        # 设定单元格的值，三种方式
        # sheet.cell(row=2, column=5).value = 99
        # sheet.cell(row=3, column=5, value=100)
        # ws['A4'] = 4  # write
        data=self.data
        data.cell(row=row,column=col).value=value
        if is_save:
            self.save_workbook()

    def write_values(self,values:(list,tuple,dict),is_save=False):
        '''
        使用list、tuple、dict批量写入数据至excel
        Args:
            values:
        Returns:
        '''
        data=self.data
        # print("values:",values)
        data.append(values)
        if is_save:
            self.save_workbook()
    # 基于ws删除一些行和一些列，注意没有备份,
    # 并且最后没有保存，因为可能ws还有后续操作
    # 必须记得最后还要保存。
    def del_ws_rows(self, rows,is_save=False):
        """基于ws删除一"""
        if not isinstance(rows,(list,tuple)):
            rows=[rows]
        rows = sorted(rows, reverse=True)  # 确保大的行数首先删除
        for row in rows:  # rowd格式如：[1,3,5],表示要删除第1、3、5共三行。
            self.data.delete_rows(row)
        if is_save:
            self.save_workbook()

    def del_ws_cols(self, cols,is_save=False):
        """基于ws删除一些列"""
        if not isinstance(cols,(list,tuple)):
            cols=[cols]
        cols = sorted(cols, reverse=True)  # 确保大的列数首先删除
        for col in cols:  # cold格式如：[2,6,10],表示要删除第2、6、10共三列
            self.data.delete_cols(col)
        if is_save:
            self.save_workbook()

    #保存
    def save_workbook(self,file_address=None):
        if file_address is None:
            file_address=self.file_address
        else:
            try:
                file_address=os.path.join(self.file_path,file_address)
            except BaseException as error:
                print("保存的文件路径还是不要乱来哦")
                raise Exception(error)
        # print(file_address)
        self.workbook.save(file_address)

    # 获取工作簿对象并返回，property将方法转换为属性
    # 获取sheets的内容
    def get_data(self,sheet_id=None):
        data=self.workbook
        if sheet_id is None:
            sheet_id=0
        else:
            sheet_id=int(sheet_id)
            # raise TypeError("sheet_id类型错误")
        tables = data[(data.sheetnames[sheet_id])]
        # 返回数据前先保存关闭
        return tables

    def get_all_sheet_name(self):
        '''获取所有sheet_name'''
        return self.workbook.sheetnames


    def get_data_for_sheet_name(self,sheet_name=None):
        sheet_names = self.get_all_sheet_name()
        if not sheet_name:
            sheet_name=sheet_names[0]
        if sheet_name not in sheet_names:
            raise sheetNameNotFoundError(f"sheetName不在定义的列表范围内,预期：{sheet_names},实际为:{sheet_name}")
        tabels=self.workbook[sheet_name]
        return tabels

    # 获取单元格的行数
    def get_lines(self,isRow=True):
        tables = self.data
        '''max_row:获取行数，max_column:获取列数
        isRow:默认获取最大行数，flase：获取最大列数'''
        if isRow:
            lines=tables.max_row
        else:
            lines = tables.max_column
        return lines

    # 获取有效的行数,过滤一行数据全是None获取空格的情况
    def get_valid_lines(self):
        valid_lines=0
        lines=self.get_lines(isRow=True)
        for i in range(1,lines+1):
            line_data=self.get_row_data(i)[0] #行数据第一个字段为空或者为空格，当做无效行
            if line_data and str(line_data).strip():
                valid_lines+=1
            else:
                break
        return valid_lines

    # 获取某一个单元格的内容
    def get_cell_value(self, row, col):
        cell_value=self.data.cell(row=row,column=col).value
        # 也可以使用：cell_value=self.data['A1'].value
        return cell_value
        # 根据行列返回表单内容

    
    # 根据对应的caseid找到对应行的内容
    def get_row_data(self, case_id):
        row_num = self.get_row_num(case_id)
        rows_data = self.get_row_values(row_num)
        return rows_data

    # 根据对应的caseid找到相应的行号
    def get_row_num(self, case_id):
        '''用例起始行为2，所以这里需要指定now初始值为2'''
        try:
            num = 2
            cols_data = self.get_col_data(1)
            cols_data=[int(i) for i in cols_data]
            lines=self.get_lines(isRow=True)
            try:
                case_id=int(case_id)
                if case_id<=lines:
                    if cols_data:
                        for col_data in cols_data:
                            if case_id == col_data:
                                return num
                            num = num + 1
                    else:
                        return None
                else:
                    print('依赖caseId不能大于用例总行数')
                    return None
            except TypeError as typeerror:
                # print(typeerror)
                return None
        except (ValueError,TypeError) as e:
            # print("excel 第一行内容不是int")
            return case_id
            
    # 根据行号找到该行内容
    def get_row_values(self, row):
        cols=self.get_lines(isRow=False)#获取最大列
        rowdata=[]
        for i in range(1,cols+1):
            cellvalue=self.data.cell(row=row,column=i).value
            rowdata.append(cellvalue)
        return rowdata

    # 获取某一列的内容
    def get_col_data(self, col=None):
        rows = self.get_lines(isRow=True)#获取最大行
        columndata=[]
        for i in range(1,rows+1):
            if col != None:
                cellvalue = self.data.cell(row=i,column=col).value
            else:
                cellvalue=self.data.cell(row=i,column=1).value
            columndata.append(cellvalue)
        return columndata
    #
    def __del__(self):
        self.workbook.close()


class operationExcelXlsxwriter(object):

    def __init__(self,file_path,file_name):
        self.file_path=file_path
        self.file_name=file_name
        if self.file_path is None:
            self.file_path = os.path.join(os.path.dirname(__file__), '../config')
            self.file_name='DataCase_ALL.xlsx'
        self.file_address = os.path.join(self.file_path, self.file_name)
            # # 类实例化时即完成对工作薄对象进行实例化，访问写入的数据不能立马生效的问题
        if not os.path.exists(self.file_path):
            os.makedirs(self.file_path)
        if  os.path.exists(self.file_address):
           os.remove(self.file_address)
        self.workbook = xlsxwriter.Workbook(self.file_address, options={'constant_memory': True})
        self.worksheet=None

    def get_all_sheet_name(self):
        return self.workbook.sheetnames

    def get_worksheet(self,sheet_name):
        sheet_names=self.get_all_sheet_name()
        if sheet_name not in sheet_names.keys():
            worksheet = self.workbook.add_worksheet(name=sheet_name)  # 默认为 Sheet1.
        else:
            worksheet=self.workbook.get_worksheet_by_name(name=sheet_name)
        return worksheet

    def write_value(self,sheet_name,row,col,value):
        worksheet=self.get_worksheet(sheet_name=sheet_name)
        worksheet.write(row, col,value)

    def write_values(self,row,col,values,sheet_name=None):
        '''

        Args:
            sheet_name:
            row: 行，从0行开始读
            col: 列，通常情况下建议为0
            values: list对象

        Returns:

        '''

        '''教程：https://blog.csdn.net/shammy_feng/article/details/124149896'''
        if not (self.worksheet or sheet_name):
            self.worksheet=self.get_worksheet(sheet_name=sheet_name)
        self.worksheet.write_row(row=row,col=col,data=values)



    def save_and_close(self):
        self.workbook.close()



class operationJson(object):
    # 初始化文件
    def __init__(self,json_path=None, file_path=None):
        if json_path is None:
            json_path = os.path.join(os.path.dirname(__file__), '../config')
        else:
            json_path=json_path
        if file_path == None:
            self.file_path = os.path.join(json_path,'Moudle_test.json')
        else:
            self.file_path = os.path.join(json_path,file_path)

        self.dict={}#存放需要的数据
        # self.data = self.read_data()
    
    def read_data(self):
        # with 语句适用于对资源进行访问的场合，确保不管使用过程中是否发生异常都会执行必要的“清理”操作，释放资源，
        # 比如文件使用后自动关闭／线程中锁的自动获取和释放
        # 读取json文件
        with open(self.file_path, encoding='utf-8') as fp:
            json_data=fp.read()
            data = json.loads(json_data)
            return data
        # 根据关键字获取数据
    
    def get_data(self, id=None):
        data=self.read_data()
        if id != None:
            return data[id]
        else:
            print('id不存在，请检查id是否填写正确')
            return None
        # 写入json
    def write_data(self, data):
        with open(self.file_path, 'w') as fp:
            fp.write(json.dumps(data,ensure_ascii=False))


if __name__ == '__main__':
    # dict1 = {'crm_course_name':{"is_refund_audit":333,'this':100}}
    # opym=OperationYaml()#'dependFieldInfo.yaml'  #'dependKeyInfo.yaml'
    # print(opym.readDataForKey('config'))
    # opx=operationExcel(file_path='../test_data/interface',file_name='maint-apiv2-api-docs11.xlsx',data_only=False,read_only=True)
    # opx=operationExcel(file_path='../test_data/interface',file_name='maint-apiv2-api-docs111.xlsx',re=False,data_only=False,read_only=False)
    opx=operationExcel(file_path='../excel',file_name='auto_genrate_api_case_allpairspy_params_obj_params_custom.xlsx',opeartion_type='read_write',data_only=False)
    opx.data=opx.get_data_for_sheet_name(sheet_name='设备管理')
    print(opx.get_col_data(col=None))
    # opx.save_workbook()
    # print(opx.get_all_sheet_name())
    # opx=operationExcelXlsxwriter(file_path='../test_data/interface',file_name='auto_genrate_api_case_allpairspy_params_obj-31.xlsx')
    # opx.xlsxwriter_write(sheet_name='666',row=1,col=1,value=['a','b','c'])
    # print(opx.get_all_sheet_name())
    # opx.save_and_close()
    # print(opx.get_all_sheet_name())
    # print(opx.get_lines())
    # print(opx.get_valid_lines())
    # print(opx.get_lines())
    # print(opx.workbook_clear())
    # print(opx.del_ws_cols(1))
    # opx.create_sheet(title="888")
    # opx.delete_sheet(sheet_name="应用管理_获取OSS信息")
    # opx.rename_sheet_name(src_sheet_name='Sheet',target_sheet_name='666',is_save=True)
    # opx.copy_worksheet('Sheet','copy_test',is_save=True)
    # print(opx.get_all_sheet_name())
    # print(opx.get_data_for_sheet_name())
    # print(opx.del_ws_rows(1,is_save=True))
    # print(opx.del_ws_cols(1,is_save=True))
    # print(opx.get_data_for_sheet_name())
    # print(opx.get_data_for_sheet_name(sheet_name='business1'))
    # print(opx.get_lines())
    # print(opx.data.max_row)
    # print(opx.data.max_column)
    # print(opx.get_valid_lines())
    # # print(opx.get_col_data(7))
    # print(opx.get_row_num(11))
